#! python3
# Format the xlsx file

import openpyxl, string
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

chars = list(string.ascii_lowercase)

def setWidthColumn (sheet, currentCell, currentCol): 
    """ Set the with for each column """

    if currentCell.value == "X":
        width = len(str(currentCell.value)) + 5
    else: 
        width = len(str(currentCell.value)) + 2

    if currentCell.value == "X" or width > sheet.column_dimensions[chars[currentCol-1]].width: 
        sheet.column_dimensions[chars[currentCol-1]].width = width

def formatData (filePath, sheetName, colStart, rowStart, data): 

    wb = openpyxl.load_workbook(filePath) 
    sheet = wb [sheetName]
    
    colNum = len(data[0])
    rowNum = len(data)

    table = chars[colStart-1].upper() + str(rowStart) + ":" + chars[colStart + colNum -1].upper() + str(rowStart + rowNum)
    print ('Applying styles to table %s' % (table))

    # Get and format cells
    for currentRow in range (rowStart, rowStart + rowNum): 
        for currentCol in range (colStart, colStart + colNum):
            titleCell = "%s%s" % (chars[currentCol-1], rowStart - 2)
            subtitleCell = "%s%s" % (chars[currentCol-1], rowStart - 1)
            currentCellName = "%s%s" % (chars[currentCol-1], currentRow)
            currentCell = sheet[currentCellName]

            titleName = str(sheet[titleCell].value).lower().strip()
            subtitleName = str(sheet[subtitleCell].value).lower().strip()

            # Set width for all +columns
            setWidthColumn (sheet, currentCell, currentCol)

            if titleName == "fecha": 
                currentCell.alignment = Alignment(horizontal='center')                   
            elif subtitleName == "público g." or subtitleName == "clientes": 
                currentCell.alignment = Alignment(horizontal='right')
                currentCell.number_format = '#,##0.00'
            elif subtitleName == "ing": 
                currentCell.font = Font (size=12, bold=True)
                currentCell.alignment = Alignment(horizontal='center')
                currentCell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type = "solid")
            elif subtitleName == "egr": 
                currentCell.font = Font (size=12, bold=True)
                currentCell.alignment = Alignment(horizontal='center')
                currentCell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type = "solid")
            elif subtitleName == "pag": 
                currentCell.font = Font (size=12, bold=True)
                currentCell.alignment = Alignment(horizontal='center')
                currentCell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type = "solid")
            elif titleName == "folio": 
                currentCell.alignment = Alignment(horizontal='center')
            elif titleName == "proveedor": 
                currentCell.alignment = Alignment(horizontal='left')
            elif subtitleName == "importe" or subtitleName == "iva" or titleName == "total":
                currentCell.alignment = Alignment(horizontal='right')
                currentCell.number_format = '#,##0.00'  
            elif titleName == "comentarios": 
                currentCell.alignment = Alignment(horizontal='left')

    # Set format to totals
    totalsRow = rowStart + rowNum 
    for currentCol in range (colStart, colStart + colNum):
        currentCellName = "%s%s" % (chars[currentCol-1], totalsRow)
        currentCell = sheet[currentCellName]
        if currentCell.value: 
            currentCell.alignment = Alignment(horizontal='right')
            currentCell.number_format = '#,##0.00'
            currentCell.font = Font(bold=True)
            currentCell.border = Border(top=Side(border_style="thin", color='FF000000'))
    wb.save (filePath)

def insertLogo (filePath, sheetName, logo, anchor, width, height): 
    """ Paste logo in each page"""
    # Open file 
    wb = openpyxl.load_workbook(filePath) 
    sheet = wb [sheetName]

    # Paste logo
    img = openpyxl.drawing.image.Image(logo)
    img.anchor = anchor
    img.width = width
    img.height = height
    sheet.add_image(img)

    # Save
    wb.save (filePath)
    print ('Added image %s' % (logo))

def formatHeaders (filePath, sheetName, colStart, rowStart, colNum, rowNum): 
    """ Set format for the headers """
    wb = openpyxl.load_workbook(filePath) 
    sheet = wb [sheetName]

    # Get and format cells
    for currentRow in range (rowStart, rowStart + rowNum): 
        for currentCol in range (colStart, colStart + colNum):
            currentCellName = "%s%s" % (chars[currentCol-1], currentRow)
            currentCell = sheet[currentCellName]

            if currentCellName == "a1" or currentCellName == "a2" or currentCellName == "j1" or currentCellName == "j2":
                currentCell.font = Font(name='Corben',
                                    size=28,
                                    bold=True,
                                    color='003399')
                currentCell.alignment = Alignment(horizontal='center')
                sheet.row_dimensions[1].height = 42
                sheet.row_dimensions[2].height = 42
            elif currentCellName == "a4" or currentCellName == "a5" or currentCellName == "j4" or currentCellName == "j5": 
                currentCell.font = Font(size=14, bold=True)
                currentCell.alignment = Alignment(horizontal='center')
            elif currentRow == 7 or currentRow == 8: 
                currentCell.font = Font(bold=True)
                currentCell.alignment = Alignment(horizontal='center', vertical='center')
                currentCell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                            right=Side(border_style="thin", color='FF000000'),
                                            top=Side(border_style="thin", color='FF000000'),
                                            bottom=Side(border_style="thin", color='FF000000'))
    wb.save (filePath)



